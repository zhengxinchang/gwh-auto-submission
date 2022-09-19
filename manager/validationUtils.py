"""各种校验"""
import os
import inspect
import re
import openpyxl
from manager.mException import BreakThisBatchProcessingException
from manager.managerGlobal import mgLog2File as mlogger
#from mException import BreakThisBatchProcessingException

class dispatherValidator():

    def checkSubject(self,subject):
        """检查邮件的subject是否合法"""
        ptn = re.compile('[0-9a-f]{8}(-[0-9a-f]{4}){3}-[0-9a-f]{12}')
        subject = subject.strip()
        if not isinstance(subject,str):
            return False
    
        if subject.startswith("NEW_SUBMISSION"):
            if subject =="NEW_SUBMISSION":
                return True
            else:
                return False 
        
        elif subject.startswith("RE_SUBMISSION:BATCHID") :
            this_uuid = subject.replace("RE_SUBMISSION:BATCHID=","")
            # if re.match('^batch_\d{4}_\d{2}\d{2}_[0-9a-f]{8}(-[0-9a-f]{4}){3}-[0-9a-f]{12}$',this_uuid):
            if re.match('^batch_\d{4}_\d{2}_\d{2}_\d{2}_[0-9a-f]{8}(-[0-9a-f]{4}){3}-[0-9a-f]{12}$',this_uuid):

                return True
            else:
                return False
        else:
            return False 

    def checkExcel(self,excelfile):
        """判断文件是否是合法excel"""
        assert os.path.exists(excelfile)
        try:
            wb2 = openpyxl.load_workbook(excelfile)
            sheetnames = wb2.sheetnames
            if len(sheetnames) == 0:
                return False
            else:
                ws =  wb2.get_sheet_by_name(sheetnames[0])
                onecell = ws['A1']
                return True
        except Exception as e:
            mm = "Can not open the excel file reason:\n{}".format(str(e))
            mlogger.error(mm)
            return(False)

    def checkEmail(self,email):
        if re.match("^(\w)+(\.\w+)*@(\w)+((\.\w+)+)$",email):
            return(True)
        else:
            return False
