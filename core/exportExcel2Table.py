#!/usr/bin/env python3
'''
对excel输出到表格模块，支持 Assembly Contact Publication三个sheet
支持对空行的判断和剔除（不完成）
后期支持对最大列的判断
这个模块应当是在excel合法性校验之后运行，则默认excel是校验通过的。
还没有定义独特的异常类型，用于邮件的发送过程。

'''
import os
import re
import logging
from collections import OrderedDict
import openpyxl
from  core.workerGlobal import gLog2File as logger
from core.workerGlobal import gConfigs 
from core.workerGlobal import gMailManager as automail
import core.eachJobGlobalVars as gvar
from core.workerGlobal import StartAndEnd

class ExcelExport():

    def __init__(self,submissionExcel):
        # logging.info("ExcelExport: Parsing Batch submisson Excel file: {}".format(submissionExcel))
        logger.info("ExcelExport: Parsing Batch submisson Excel file: {}".format(submissionExcel))
        if not os.path.exists(submissionExcel):
            logger.error("config file {} does not exists".format(submissionExcel))
            exit(1)        

        self.wb = openpyxl.load_workbook(submissionExcel)

    def exportAssemblySheet(self,filename=None):
        sheet_list  = self.wb.sheetnames
        if "Assembly" not in sheet_list:
            raise Exception("Sheet Assembly not in Excel")
        sheet_data = self.wb['Assembly']
        if filename != None :
            count = 0
            with open(filename, 'w') as fout :
                
                for r in sheet_data.iter_rows():
                    # print(r)
                    count += 1
                    if count < 20 :
                        continue
                    # print(count)
                    if not self.__onlyNone(r):
                        # print("\t".join([str(cell.value) for cell in r]) + "\n")
                        fout.write(self.__joinLine(r) + "\n")    

        else:
            out = []
            count = 0
            for r in sheet_data.iter_rows():
                count += 1
                if count < 20:
                    continue
                if not self.__onlyNone(r):
                    out.append(self.__joinLineList(r))    
            return(out)

    def exportContactSheet(self,filename=None):

        sheet_list  = self.wb.sheetnames
        if "Contact" not in sheet_list:
            raise Exception("Sheet Contact not in Excel")
        sheet_data = self.wb['Contact']
        if filename != None :
            count = 0
            with open(filename, 'w') as fout :
                
                for r in sheet_data.iter_rows():
                    count += 1
                    if count < 10 :
                        continue
                    if not self.__onlyNone(r):
                        # print("\t".join([str(cell.value) for cell in r]) + "\n")
                        fout.write(self.__joinLine(r) + "\n")    
        else:
            out = []
            count = 0
            for r in sheet_data.iter_rows():
                count += 1
                if count < 10:
                    continue
                if not self.__onlyNone(r):
                    out.append(self.__joinLineList(r))     
            # print(out)
            return(out)

    def exportPublicationSheet(self,filename=None):
        sheet_list  = self.wb.sheetnames
        if "Publication" not in sheet_list:
            raise Exception("Sheet Publication not in Excel")
        sheet_data = self.wb['Publication']
        if filename != None :
            count = 0
            with open(filename, 'w') as fout :
                for r in sheet_data.iter_rows():
                    count += 1
                    if count < 12 :
                        continue
                    if not self.__onlyNone(r):
                        # print("\t".join([str(cell.value) for cell in r]) + "\n")
                        fout.write( self.__joinLine(r) + "\n")    
        else:
            out = []
            count = 0
            for r in sheet_data.iter_rows():
                count += 1
                if count < 12:
                    continue
                if not self.__onlyNone(r):
                    out.append(self.__joinLineList(r))    
            return(out)

    def __joinLine(self,row):
        outLine = []
        for k in row:
            # print(k.value)
            if k.value == None:
                outLine.append("")
            else:
                outLine.append(str(k.value))
        return("\t".join(outLine))
        
    def __joinLineList(self,row):
        outLine = []
        for k in row:
            # print(k.value)
            if k.value == None:
                outLine.append("")
            else:
                outLine.append(str(k.value))  
        return(outLine)

    def __onlyNone(self,list_row):
        flag =  True
        # for k in list_row:
        #     if None != k.value and (not re.match("^ +$",str(k.value))):
        #         flag = False
        
        for d,k in enumerate(list_row):
            if d <1 :
                continue
            # print(k)
            if None != k.value and (not re.match("^ +$",str(k.value))):
                flag = False 

        return(flag)

class Excel2Objects():
    """
    excel 三个表格导出为三个列表，每个列表中元素是一个字典，对应的是一条记录（excel中一行）
    """
    def __init__(self,submissionExcel):

        logger.info("Excel2Objects: Parsing Batch submisson Excel file: {}".format(submissionExcel))
        if not os.path.exists(submissionExcel):
            logging.error("config file {} does not exists".format(submissionExcel))
            automail.mail2manager("config file {} does not exists".format(submissionExcel))
            StartAndEnd.end() 

        self.wb = openpyxl.load_workbook(submissionExcel)
        self.excel_params = gConfigs.get("excel","EXCEL_SHEETS_AND_COLUMNS")
        # {Contact:[第几行是真实数据，一共的列数]}
        self.excel_params_dict = {i.split(":")[0]:[int(i.split(":")[1]),int(i.split(":")[2])] for i in self.excel_params.split(",")}

    def getContactObjsList(self,):
        sheet_list  = self.wb.sheetnames
        if "Contact" not in sheet_list:
            gvar.Message.addExcelMessage("Contact sheet not found!")
            return(None)
        sheet_data = self.wb['Contact']
        out = []
        count = 0
        for r in sheet_data.iter_rows():
            count += 1
            if count < self.excel_params_dict["Contact"][0]:
                continue
            line_dict = {}
            if not self.__onlyNone(r): 
                for i,c in enumerate(r):
                    if i > self.excel_params_dict["Contact"][1]:
                        continue
                    elif i == 0 :                         
                        line_dict["CID"] =self.__cleanValue(c.value)
                    elif i == 1 :
                        line_dict["First name"]=self.__cleanValue(c.value)
                    elif i == 2 :
                        line_dict["Middle name"]=self.__cleanValue(c.value)
                    elif i == 3 :
                        line_dict["Last name"]=self.__cleanValue(c.value)
                    elif i == 4 :
                        line_dict["Email"]=self.__cleanValue(c.value)
                    elif i == 5 :
                        line_dict["Email (secondary)"]=self.__cleanValue(c.value)   
                    elif i == 6 :    
                        line_dict["Organization"]=self.__cleanValue(c.value)
                    elif i == 7 :    
                        line_dict["Department"]=self.__cleanValue(c.value)
                    elif i == 8 :    
                        line_dict["Phone"]=self.__cleanValue(c.value)
                    elif i == 9 :    
                        line_dict["Fax"]=self.__cleanValue(c.value)
                    elif i == 10 :    
                        line_dict["Street"]=self.__cleanValue(c.value)
                    elif i == 11 :    
                        line_dict["City"]=self.__cleanValue(c.value)
                    elif i == 12 :    
                        line_dict["State/Province"]=self.__cleanValue(c.value)
                    elif i == 13 :    
                        line_dict["Postal code"]=self.__cleanValue(c.value)
                    elif i == 14 :    
                        line_dict["Country/Region"]=self.__cleanValue(c.value)
                    elif i == 15 :    
                        line_dict["BIGD_account"]=self.__cleanValue(c.value)

                out.append(line_dict)
        return(out)

    def getPublicationObjsList(self,):
        sheet_list  = self.wb.sheetnames
        if "Publication" not in sheet_list:
            gvar.Message.addExcelMessage("Publication sheet not found!")
            return None
        sheet_data = self.wb['Publication']
        out = []
        count = 0
        for r in sheet_data.iter_rows():
            count += 1
            if count < self.excel_params_dict["Publication"][0]:
                continue
            line_dict = {}
            if not self.__onlyNone(r): 
                for i,c in enumerate(r):
                    if i > self.excel_params_dict["Publication"][1]:
                        continue
                    elif i == 0 :                         
                        line_dict["PID"] =self.__cleanValue(c.value)
                    elif i == 1 :
                        line_dict["Status"]=self.__cleanValue(c.value)
                    elif i == 2 :
                        line_dict["PubMed ID"]=self.__cleanValue(c.value)
                    elif i == 3 :
                        line_dict["Paper URL"]=self.__cleanValue(c.value)
                    elif i == 4 :
                        line_dict["Title"]=self.__cleanValue(c.value)
                    elif i == 5 :
                        line_dict["Authors"]=self.__cleanValue(c.value)   
                    elif i == 6 :    
                        line_dict["Journal"]=self.__cleanValue(c.value)
                    elif i == 7 :    
                        line_dict["Year"]=self.__cleanValue(c.value)
                    elif i == 8 :    
                        line_dict["Volume"]=self.__cleanValue(c.value)
                    elif i == 9 :    
                        line_dict["Issue"]=self.__cleanValue(c.value)
                    elif i == 10 :    
                        line_dict["Pages from"]=self.__cleanValue(c.value)
                    elif i == 11 :    
                        line_dict["Pages to"]=self.__cleanValue(c.value)

                out.append(line_dict)
        return(out)

    def getAssemblyObjsList(self,):
        sheet_list  = self.wb.sheetnames
        if "Assembly" not in sheet_list:
            gvar.Message.addExcelMessage("Assembly sheet not found!")
            return(None)

        sheet_data = self.wb['Assembly']
        out = []
        count = 0
        for r in sheet_data.iter_rows():
            count += 1
            if count < self.excel_params_dict["Assembly"][0]:
                continue
            line_dict = {}
            if not self.__onlyNone(r): 
                for i,c in enumerate(r):
                    if i > self.excel_params_dict["Assembly"][1]:
                        continue
                    if i == 0 :
                        line_dict["GID"]=self.__cleanValue(c.value)
                    elif i == 1 :
                        line_dict["CID"]=self.__cleanValue(c.value)
                    elif i == 2 :
                        line_dict["PID"]=self.__cleanValue(c.value)
                    elif i == 3 :
                        line_dict["Submission title"]=self.__cleanValue(c.value)
                    elif i == 4 :
                        line_dict["Sequence authors name"]=self.__cleanValue(c.value)
                    elif i == 5 :
                        line_dict["Contact authors name"]=self.__cleanValue(c.value)
                    elif i == 6 :
                        line_dict["Biosample Accession"]=self.__cleanValue(c.value)
                    elif i == 7 :
                        line_dict["Biosproject Accession"]=self.__cleanValue(c.value)
                    elif i == 8 :
                        line_dict["Release date"]=self.__cleanValue(c.value)
                    elif i == 9 :
                        line_dict["Reference assembly name or accession"]=self.__cleanValue(c.value)
                    elif i == 10 :
                        line_dict["Existing genome accession"]=self.__cleanValue(c.value)
                    elif i == 11 :
                        line_dict["Assembly name"]=self.__cleanValue(c.value)
                    elif i == 12 :
                        line_dict["Assembly method"]=self.__cleanValue(c.value)
                    elif i == 13 :
                        line_dict["Program Version or release date"]=self.__cleanValue(c.value)
                    elif i == 14 :
                        line_dict["Sequencing technology"]=self.__cleanValue(c.value)
                    elif i == 15 :
                        line_dict["Genome coverage"]=self.__cleanValue(c.value)
                    elif i == 16 :
                        line_dict["Sequence Contamination QC"]=self.__cleanValue(c.value)
                    elif i == 17 :
                        line_dict["Genome composition"]=self.__cleanValue(c.value)
                    elif i == 18 :
                        line_dict["Assembly level"]=self.__cleanValue(c.value)
                    elif i == 19 :
                        line_dict["Genome sequence file name"]=self.__cleanValue(c.value)
                    elif i == 20 :
                        line_dict["Genome sequence file MD5 code"]=self.__cleanValue(c.value)
                    elif i == 21 :
                        line_dict["Genome annotation file name"]=self.__cleanValue(c.value)
                    elif i == 22 :
                        line_dict["Genome annotation file MD5 code"]=self.__cleanValue(c.value)
                    elif i == 23 :
                        line_dict["Contig to scaffold AGP file name"]=self.__cleanValue(c.value)
                    elif i == 24 :
                        line_dict["Contig to scaffold AGP file MD5 code"]=self.__cleanValue(c.value)
                    elif i == 25 :
                        line_dict["Other AGP file name"]=self.__cleanValue(c.value)
                    elif i == 26 :
                        line_dict["Other AGP file MD5 code"]=self.__cleanValue(c.value)
                    elif i == 27 :
                        line_dict["Chromosome assignment file name"]=self.__cleanValue(c.value)
                    elif i == 28 :
                        line_dict["Chromosome assignment file MD5 code"]=self.__cleanValue(c.value)
                    elif i == 29 :
                        line_dict["Plasmid assignment file name"]=self.__cleanValue(c.value)
                    elif i == 30 :
                        line_dict["Plasmid assignment file MD5 code"]=self.__cleanValue(c.value)
                    elif i == 31 :
                        line_dict["Organella assignment file name"]=self.__cleanValue(c.value)
                    elif i == 32 :
                        line_dict["Organella assignment file MD5 code"]=self.__cleanValue(c.value)

                out.append(line_dict)
        # print(out)
        return(out)

    def __onlyNone(self,list_row):
        flag =  True

        for d,k in enumerate(list_row):
            if d <1 :
                continue
            # print(k)
            if None != k.value and (not re.match("^ +$",str(k.value))):
                flag = False 

        return(flag)
    def __cleanValue_doubleQuotas(self,cvalue):
        if cvalue == None:
            return("NULL")
        elif type(cvalue) ==  str:
            a = str(cvalue).strip()
            if a.strip() == "":
                a = "NULL"
                return(a)
            else:
                # a = "'" + a.replace("'"," ") + "'" #防止字符串中出现 ' 符号，影响sql插入
                a = '"' + a.replace('"'," ") + '"' #防止字符串中出现 " 符号，影响sql插入
                return(a)
        else:
            return(cvalue)

    def __cleanValue(self,cvalue):
        if cvalue == None:
            return("NULL")
        # elif type(cvalue) ==  str:
        #     a = str(cvalue).strip()
        #     if a.strip() == "":
        #         a = "NULL"
        #         return(a)
        #     else:
        #         a = "'" + a.replace("'"," ") + "'" #防止字符串中出现 ' 符号，影响sql插入
        #         # a = '"' + a.replace('"'," ") + '"' #防止字符串中出现 " 符号，影响sql插入
        #         return(a)
        # else:
        #     return(str(cvalue))
        else:
            return(cvalue)
        
                        

if __name__ == "__main__":
    


    excel_exp = ExcelExport("/mnt/e/projects/NGDC/GWH/autoBatchSubmissionPipeline/excel/GWH-batchsubmission-Chinese-small.xlsx")

    excel_obj = Excel2Objects("/mnt/e/projects/NGDC/GWH/autoBatchSubmissionPipeline/excel/GWH-batchsubmission-Chinese-small.xlsx",conf)

    # print(excel_obj.getContactObjsList())

    # print(excel_obj.getPublicationObjsList())
    
    # print(excel_obj.getAssemblyObjsList())
    # print(excel_exp.exportAssemblySheet(filename="./assemblysheet.txt"))
    # print(excel_exp.exportContactSheet(filename="./contactsheet.txt"))
    # print(excel_exp.exportPublicationSheet(filename="./publicationsheet.txt"))